#!/usr/bin/env python

"""
Read the rangeland monitoring observational spreadsheets (microsoft excel) and extracts out the major cover fractions and tree basal area along with the site locations, date of data capture and site name. The script will generate a unique identifier for each observatial spread sheet using the site name and date visited, enabling multiple years to be processed. 

Author: Grant Staben
email: grant.staben@nt.gov.au
Date: 07/12/2021
Version: 1.0

Adapted from original scripts written by Robert McGregor (2017, Complete_works_v13_Fraction-Copy1.py) and Daniel McIntyre (26/11/2020, Calculate FPC.ipynb) which extracted woody FPC, PPC, CC and other parameters from csv files or spreadsheets. 


###############################################################################################

MIT License

Copyright (c) 2021 Grant Staben

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

###############################################################################################

Parameters: 
-----------


indir : str 
            is a string containing the path to the directory containing the RM observational spreadsheets to be processed.

csv : str
            is a string containing the name of the csv file containing the results along with the path of the directory to save it into.


"""


# Import python modules
import pandas as pd
import os
import argparse
import sys
from openpyxl import load_workbook


# command arguments
def getCmdargs():

    p = argparse.ArgumentParser(description="""This script reads in rangeland monitoring observational spreadsheets and extract out the major cover fractions along with the site locations, date of data capture and site name. It concatenates the results for each site into one csv file to be used for further analysis.""")

    p.add_argument("-d","--indir", help="Path to the directory containing the observational spreadsheets")
 
    p.add_argument("-o","--csv", help="Path and name of the output csv file containing the results")
    
    cmdargs = p.parse_args()
    
    if cmdargs.indir is None:

        p.print_help()

        sys.exit()

    return cmdargs


def FPC(row):
    '''
    Function FPC will count the number of intercepts classified as woody green foliage.
    '''
    
    # above green
    if row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK':
        val = 1
              
    # in crown   
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1
           
    # not in crown (blank)
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 0
        
    else:
        val = 0
        
    return val


def PPC(row):
    '''
    Function PPC will count the number of intercepts classified as woody green and non-photosynthetic foliage/branches to calculate woody
    plant projective cover.
    '''
    # above green
    if row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK':
        val = 1
     
     # above brown  
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK':
        val = 1        
                  
    # above dead   
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK':
        val = 1 
            
    # in crown   
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1

    # not in crown (blank)
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1  

    else:
        val = 0
        
    return val


def CC(row):
    '''
    Function PPC will count the number of intercepts classified as woody green and non-photosynthetic foliage/branches and falling within
    the crown of trees to calculate woody canopy cover.
    '''
    # above green
    if row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK':
        val = 1
  
     # above brown  
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK':
        val = 1        
               
    # above dead   
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK':
        val = 1 
           
    # in crown   
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK':
        val = 1 
         
    # not in crown (blank)  
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1    
        
    else:
        val = 0
        
    return val


def OB(row):
    '''
    Function OB will calculate the number of "ABOVE" branch intercepts which may have occluded woody green leaf intercepts - this is used
    take into acount the probability that green leaf intercepts were likely to have been obscured due to the way the intercepts are
    collected (bottom-up) where as the satellite sensor is top-down. 
    The equation used to calculate woody fpc follows Armston et al. 2009  
    '''  
    # above brown
    if row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK':
        val = 1 
    
    else:
        val = 0
        
    return val


def pv(row):
    """ 
    pv function calculates the total intercepts of photosynthetic green vegetation. This is based on the first intercept 
    that the satellite sensor will see when look down from the top.
    """
    ############################################# above green - below green and ground layer options ###################################
    
    if row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1
       
    #########################################  above green - below brown - ground layer options #####################################
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1 
    
    ############################################# above green - below dead - ground layers #####################################    
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1        
    ############################################# above green - sub-shrub - ground layers #####################################    
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1    
   ############################################# above green - None - ground layers #####################################        
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1   
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1   
    elif row['ABOVE'] == 'ABOVE - GREEN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1          
   ############################################# In Crown sky - Below Green - ground layers #####################################   
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1     
    ############################################# Blank - Below Green - ground layers #####################################   
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1  
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1         
    ############################################# Blank - subshrub-grey - ground layers #####################################   
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1        
    ############################################# In Crown sky - subshrub-grey - ground layers #####################################   
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1    
   ############################################# In Crown sky - Blank - Green ground intercepts ########################## 
   
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1
 
    ############################################# Blank - Blank - Green ground intercepts ########################## 
   
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1      
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1
    
    else:
        val = 0
        
    return val 
    
    
def npv(row):
    """ npv function calculates the total intercepts of non photosynthetic green vegetation. 
        This is based on the first intercept that the satellite sensor will see when look down from the top"""
    
    ############################################# above brown - below green and ground layer options ###################################
    if row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ROCK':
        val = 1   
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1        
        
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1
        
    #########################################  above brown - below brown - ground layer options #####################################
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
        
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1     
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1     
    
    ############################################# above brown - below dead - ground layers #####################################    
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
        
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1
        
    ############################################# above brown - subshrub -grey - ground layers #####################################    
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
        
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1     
   ############################################# above brown - None - ground layers #####################################        
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1  
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1         
    elif row['ABOVE'] == 'ABOVE - BROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1          
        
  ############################################# above dead - below green and ground layer options ###################################
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1       
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - GREEN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1         
        
    #########################################  above dead - below brown - ground layer options #####################################
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1 
                
    #########################################  above dead - subshrub - grey - ground layer options #####################################
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'SUBSHRUB - GREY' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1 
    
    ############################################# above dead - below dead - ground layers #####################################    
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
        
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1
    
   ############################################# above dead - None - ground layers #####################################        
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
        
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1
    elif row['ABOVE'] == 'ABOVE - DEAD' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1 
 
    
   ############################################# In Crown sky - Below brown - ground layers #####################################   
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
        
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1     
    ############################################# Blank - Below Brown - ground layers #####################################   
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1  
        
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1  
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1  
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - BROWN' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1   
    ############################################# In Crown sky - Below dead - ground layers #####################################   
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1
        
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1
    
    ############################################# Blank - Below dead - ground layers #####################################   
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'LITTER':
        val = 1    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ROCK':
        val = 1         
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1        
        
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'ASH':
        val = 1      
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1  
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BELOW - DEAD' and row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1


   ############################################# In Crown sky - Blank - DEAD ground intercepts ########################## 
   
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1 
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1     
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'LITTER':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1

    ############################################# Blank - Blank - DEAD ground intercepts ########################## 
   
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1      
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1     
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1        
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'LITTER':
        val = 1     
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1 
    
    else:
        val = 0
        
    return val 
        
    
def bg(row):
    
    """ npv function calculates the total intercepts in the bare ground category. 
        This is based on the first intercept that the satellite sensor will see when look down from the top"""
    
    ############################################# In Crown sky - Blank - bare ground intercepts #########################
    
    if row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1      
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1     
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ROCK':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ASH':
        val = 1        
    elif row['ABOVE'] == 'ABOVE - IN CROWN' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1
 
    ############################################# Blank - Blank - bare ground intercepts ########################## 
   
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'BARE GROUND':
        val = 1      
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'GRAVEL':
        val = 1     
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ROCK':
        val = 1        
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'ASH':
        val = 1     
    elif row['ABOVE'] == 'BLANK' and row['BELOW'] == 'BLANK' and row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1     
      
    else:
        val = 0
        
    return val 
    

def groundGreen(row):# identify pv fraction components
    
    if row['GROUND LAYER'] == 'GREEN ANNUAL GRASS':
        val = 1     
    elif row['GROUND LAYER'] == 'GREEN PERENNIAL GRASS':
        val = 1 
    elif row['GROUND LAYER'] == 'GREEN ANNUAL FORB / HERB':
        val = 1 
    elif row['GROUND LAYER'] == 'GREEN PERENNIAL FORB / HERB':
        val = 1 
    elif row['GROUND LAYER'] == 'GREEN PLANT':
        val = 1 
    
    else:
        val = 0
        
    return val


def groundNPV(row):# identify npv fraction components
         
    if row['GROUND LAYER'] == 'LITTER':
        val = 1       
    elif row['GROUND LAYER'] == 'DEAD ANNUAL GRASS':
        val = 1
    elif row['GROUND LAYER'] == 'DEAD PERENNIAL GRASS':
        val = 1
    elif row['GROUND LAYER'] == 'DEAD ANNUAL FORB / HERB':
        val = 1
    elif row['GROUND LAYER'] == 'DEAD PERENNIAL FORB / HERB':
        val = 1
    elif row['GROUND LAYER'] == 'DEAD PLANT':
        val = 1       
    
    else:
        val = 0
        
    return val


def groundBare(row):# identify bare ground fraction components
         
    if row['GROUND LAYER'] == 'BARE GROUND':
        val = 1 
    elif row['GROUND LAYER'] == 'GRAVEL':
        val = 1
    elif row['GROUND LAYER'] == 'ROCK':
        val = 1
    elif row['GROUND LAYER'] == 'ASH':
        val = 1
    elif row['GROUND LAYER'] == 'CRYPTOGRAM':
        val = 1
    
    else:
        val = 0
              
    return val


def CoverIndices(appended_data):

    # identify the site name variables so that each site can be grouped and run through a for loop.
    unique_id = appended_data['uid'].unique()
        
    results_fpc = [] # create an empty list to store the output rows into

    for key in unique_id:
        # utilise the unique identifiers to run through a For loop
        select_df = appended_data[(appended_data.uid == key)]
    
        # get the sum for each column
        fpc_sum = select_df['fpc'].sum()
        ppc_sum = select_df['ppc'].sum()
        cc_sum = select_df['cc'].sum() 
        ob_sum = select_df['ob'].sum()
        pvg_sum = select_df['pvg'].sum()
        npvg_sum = select_df['npvg'].sum()
        bgg_sum = select_df['bgg'].sum()
        pv_sum = select_df['pv'].sum()
        npv_sum = select_df['npv'].sum()
        bg_sum = select_df['bg'].sum()
           
        # calculate the fpc, ppc and cc for each site          
        fpc_1 = fpc_sum/(300-ob_sum)*100
        ppc_1 = ppc_sum/300*100
        cc_1 = cc_sum/300*100
        pvg_1 = pvg_sum/300*100
        npvg_1 = npvg_sum/300*100
        bgg_1 = bgg_sum/300*100
        pv_1 = pv_sum/300*100
        npv_1 = npv_sum/300*100
        bg_1 = bg_sum/300*100
        
        # output the three variables and the key (Site) as a list
        new_row = [key, fpc_1, ppc_1, cc_1, pvg_1, npvg_1, bgg_1, pv_1, npv_1, bg_1]
        # append the list as a list of lists called new row which was created earlier.
        results_fpc.append(new_row)
    
    calculations_df = pd.DataFrame(results_fpc)
    calculations_df.columns = ["uid", "FPC", "PPC", "CC", "PVg", "NPVg", "BGg", "PV", "NPV", "BG"] 
    
    return calculations_df


def mainRoutine():

    # read in the command arguments
    cmdargs = getCmdargs()
    
    # path to the directory containing the observational spreadsheets
    path = cmdargs.indir
    # path and name of the output results
    results_csv = cmdargs.csv

    appended_data = []
    
    for wb in os.listdir(path):
        site_name = wb[-11:-5]
        full_path = path + wb
        print ('observational spreadsheet being processed: ', full_path)
        
        # read in the observational work sheet and extract out the site name and visit date to use as the uid to match the site details
        # and fractional calculations. 
        book = load_workbook(filename = full_path, data_only=True)
        sheet_ranges2 = book['Step 2 - Visit Details']
        site = (sheet_ranges2['B5'].value)
        date = (sheet_ranges2['B6'].value)
        uid = str(site) + '_' +  str(date)
        print (uid)
        print ('--------------------------------------------------')
        # read in the three transect tabs from the spreadsheet, and convert them to a concatenated dataframe
        df = pd.concat(pd.read_excel(full_path, sheet_name=[3, 4, 5], header=None, index_col=None, usecols = [1, 2, 3], skiprows = [0, 1,2], nrows=100))
        df.columns = ["GROUND LAYER", "BELOW", "ABOVE"]
            
        # drop the index columns
        df.reset_index(drop=True, inplace=True)
        #df['Site'] = site_name
        df['uid'] = uid
        appended_data.append(df)

    
    appended_data = pd.concat(appended_data, sort=True)

    
    '''Call upon the fraction calculation functions and append the variables calculations to the DataFrame'''
    appended_data['fpc'] = appended_data.apply(FPC,axis=1)
    appended_data['ppc'] = appended_data.apply(PPC,axis=1)
    appended_data['cc'] = appended_data.apply(CC,axis=1)
    appended_data['ob'] = appended_data.apply(OB,axis=1)
    appended_data['pvg'] = appended_data.apply(groundGreen,axis=1)
    appended_data['npvg'] = appended_data.apply(groundNPV,axis=1)
    appended_data['bgg'] = appended_data.apply(groundBare, axis=1)
    appended_data['pv'] = appended_data.apply(pv,axis=1)
    appended_data['npv'] = appended_data.apply(npv,axis=1)
    appended_data['bg'] = appended_data.apply(bg, axis=1)
        
    calculations_df = CoverIndices(appended_data)
    
    """
    Extract property name, site, date, lat and long for centre and the north offset from spreadsheet
    the unique id (uid) variable is created from the site and date to take into account sites with multiple visit dates.
     
    """
    appended_data = []

    for wb in os.listdir(path):

        # load the individual observational spread sheet 
        book = load_workbook(filename = path + wb, data_only=True)
        
        # read in sheet 1 to extract out the station name and lat and long coords for the centre and north offset picket
        sheet_ranges1 = book['Step 1 - Site Establishment']
        
        station = (sheet_ranges1['B6'].value)
        c_lat = (sheet_ranges1['B15'].value)
        c_lon = (sheet_ranges1['B16'].value)
        no_lat = (sheet_ranges1['B13'].value)
        no_lon = (sheet_ranges1['B14'].value)
               
        #read in sheet 2 to get the site name and visit date
        sheet_ranges2 = book['Step 2 - Visit Details']
           
        site = (sheet_ranges2['B5'].value)
        date = (sheet_ranges2['B6'].value)
        uid = str(site) + '_' + str(date)
        
        # read in the sheet to extract out the basal area data
        sheet_ranges3 = book['Step 5 - Basal Sweeps - Table 2']
        
        ba_trees = (sheet_ranges3['B17'].value) 
        ba_shrubs = (sheet_ranges3['B18'].value) 
        ba_total = (sheet_ranges3['B19'].value)
        
        # append the individual results to export to a pandas df 
        appended_data.append([station, site, date, c_lat, c_lon, no_lat, no_lon, uid, ba_trees, ba_shrubs, ba_total])
    
    attribute_df = pd.DataFrame(appended_data, columns=['Station', 'Site', 'Date', 'C_Lat', 'C_Lon', 'NO_Lat', 'NO_Lon','uid', 'ba_trees', 'ba_shrubs', 'ba_total'])
        
    # join attribute and calculation dataframes using 'uid' as index
    final_df = attribute_df.join(calculations_df.set_index('uid'), on='uid')
    
    final_df.to_csv(results_csv)

if __name__ == "__main__":
    mainRoutine()